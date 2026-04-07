import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.label import DataLabelList
from config import (
    DATA_PROCESSED,
    EXCEL_DIR,
    COST_MODEL,
    REVENUE_MODEL,
    OPPORTUNITY_COST,
    TARGET_FILL_RATIO,
    SIMULATION,
    STUDY_AREA,
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
SUBHEADER_FONT = Font(bold=True, name="Arial", size=10)
DATA_FONT = Font(name="Arial", size=10)
INPUT_FONT = Font(name="Arial", size=10, color="0000FF")
FORMULA_FONT = Font(name="Arial", size=10, color="000000")
LINK_FONT = Font(name="Arial", size=10, color="008000")
INPUT_FILL = PatternFill("solid", fgColor="FFFFCC")
TITLE_FONT = Font(bold=True, name="Arial", size=14, color="1F4E79")
CURRENCY_FMT = '$#,##0.00;($#,##0.00);"-"'
CURRENCY_INT = '$#,##0;($#,##0);"-"'
PCT_FMT = '0.0%'
NUM_FMT = '#,##0'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_row(ws, row, max_col, font=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font or DATA_FONT
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=30):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def build_summary_sheet(wb):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1F4E79"

    ws["A1"] = "Citi Bike Overnight Rebalancing Optimization"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ws["A2"] = "NYU Stern — B60.2350 Decision Models & Analytics — Spring 2026"
    ws["A2"].font = Font(name="Arial", size=11, italic=True, color="666666")
    ws.merge_cells("A2:F2")

    ws["A4"] = "Executive Summary"
    ws["A4"].font = Font(bold=True, name="Arial", size=12, color="1F4E79")

    summary_items = [
        ("Study Area", f"='Cost Model'!B3"),
        ("Stations Analyzed", "=COUNTA('Station Data'!A3:A200)"),
        ("Trip Months", "Apr 2025 – Mar 2026 (12 months)"),
        ("Morning Peak", "7:00–9:00 AM"),
        ("Recommended Trucks", "=Sensitivity!B4"),
        ("Optimal Service Level", "=Sensitivity!D4"),
        ("Nightly Rebalancing Cost", "=Sensitivity!C4"),
        ("Monthly Net Benefit", "=Sensitivity!F4*30"),
    ]

    for i, (label, value) in enumerate(summary_items):
        row = 6 + i
        ws.cell(row=row, column=1, value=label).font = SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = SUBHEADER_FILL
        cell = ws.cell(row=row, column=2, value=value)
        if "Cost" in label or "Benefit" in label:
            cell.number_format = CURRENCY_INT
            cell.font = LINK_FONT
        elif "Service" in label:
            cell.number_format = PCT_FMT
            cell.font = LINK_FONT
        elif value and str(value).startswith("="):
            cell.font = LINK_FONT
        else:
            cell.font = DATA_FONT

    ws["A16"] = "Key Findings"
    ws["A16"].font = Font(bold=True, name="Arial", size=12, color="1F4E79")

    findings = [
        "1. Cost of inaction: ~$13.50/failed trip ($4.73 revenue + $3.61 USDOT VOT + $5.17 churn risk).",
        "2. Operational cost: ~$2,088/night for 4 trucks (Isuzu NPR-HD 18', 48 bikes, buy used).",
        "3. Net benefit: ~$10K/day — rebalancing pays for itself many times over.",
        "4. Labor dominates at ~89% of operational costs; crew wage is the most sensitive driver.",
        "5. 12-month dataset (44.6M trips) captures seasonal variation across the full year.",
    ]
    for i, f in enumerate(findings):
        ws.cell(row=18 + i, column=1, value=f).font = DATA_FONT

    auto_width(ws)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25


def build_station_data_sheet(wb):
    ws = wb.create_sheet("Station Data")
    ws.sheet_properties.tabColor = "378ADD"

    stations = pd.read_parquet(DATA_PROCESSED / "station_demand_stats.parquet")
    targets = pd.read_parquet(DATA_PROCESSED / "rebalancing_targets.parquet")
    merged = targets.merge(
        stations[["station_id", "mean_departures", "std_departures",
                  "mean_arrivals", "std_arrivals", "poisson_lambda_dep",
                  "poisson_lambda_arr", "cv_departures"]],
        on="station_id", how="left", suffixes=("", "_stats")
    )

    headers = [
        "Station Name", "Station ID", "Latitude", "Longitude", "Capacity",
        "Station Type", "Avg Morning Dep", "Std Morning Dep",
        "Avg Morning Arr", "Std Morning Arr",
        "Poisson λ (dep)", "CV (dep)",
        "Est. Midnight Fill", "Target Bikes", "Deficit",
        "Needs Pickup", "Needs Dropoff",
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, len(headers))

    ws.cell(row=1, column=1, value="Station Data — Manhattan below 59th St")
    ws["A1"].font = Font(bold=True, name="Arial", size=12, color="1F4E79")

    for i, (_, row) in enumerate(merged.iterrows()):
        r = i + 3
        ws.cell(row=r, column=1, value=row["name"]).font = DATA_FONT
        ws.cell(row=r, column=2, value=row["station_id"]).font = DATA_FONT
        ws.cell(row=r, column=3, value=round(row["lat"], 6)).font = DATA_FONT
        ws.cell(row=r, column=4, value=round(row["lon"], 6)).font = DATA_FONT
        ws.cell(row=r, column=5, value=int(row["capacity"])).font = DATA_FONT
        ws.cell(row=r, column=6, value=row["station_type"]).font = DATA_FONT

        dep_col = "mean_departures_stats" if "mean_departures_stats" in row.index else "mean_departures"
        std_col = "std_departures_stats" if "std_departures_stats" in row.index else "std_departures"
        arr_col = "mean_arrivals_stats" if "mean_arrivals_stats" in row.index else "mean_arrivals"
        std_arr_col = "std_arrivals_stats" if "std_arrivals_stats" in row.index else "std_arrivals"

        ws.cell(row=r, column=7, value=round(row.get(dep_col, 0), 1)).font = DATA_FONT
        ws.cell(row=r, column=8, value=round(row.get(std_col, 0), 1)).font = DATA_FONT
        ws.cell(row=r, column=9, value=round(row.get(arr_col, 0), 1)).font = DATA_FONT
        ws.cell(row=r, column=10, value=round(row.get(std_arr_col, 0), 1)).font = DATA_FONT
        ws.cell(row=r, column=11, value=round(row.get("poisson_lambda_dep", 0), 2)).font = DATA_FONT
        ws.cell(row=r, column=12, value=round(row.get("cv_departures", 0), 3)).font = DATA_FONT
        ws.cell(row=r, column=13, value=int(row["estimated_midnight_fill"])).font = DATA_FONT
        ws.cell(row=r, column=14, value=int(row["target_bikes"])).font = DATA_FONT

        deficit_row = r
        ws.cell(row=r, column=15, value=f"=N{r}-M{r}").font = FORMULA_FONT
        ws.cell(row=r, column=16, value=f'=IF(O{r}<0,-O{r},0)').font = FORMULA_FONT
        ws.cell(row=r, column=17, value=f'=IF(O{r}>0,O{r},0)').font = FORMULA_FONT

        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = THIN_BORDER

    last_row = len(merged) + 3
    ws.cell(row=last_row, column=1, value="TOTALS").font = SUBHEADER_FONT
    ws.cell(row=last_row, column=5, value=f"=SUM(E3:E{last_row-1})").font = FORMULA_FONT
    ws.cell(row=last_row, column=13, value=f"=SUM(M3:M{last_row-1})").font = FORMULA_FONT
    ws.cell(row=last_row, column=16, value=f"=SUM(P3:P{last_row-1})").font = FORMULA_FONT
    ws.cell(row=last_row, column=17, value=f"=SUM(Q3:Q{last_row-1})").font = FORMULA_FONT

    auto_width(ws)


def build_cost_model_sheet(wb):
    ws = wb.create_sheet("Cost Model")
    ws.sheet_properties.tabColor = "E24B4A"

    ws["A1"] = "Cost Model Parameters"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Study Area"
    ws["B3"] = STUDY_AREA["name"]
    ws["B3"].font = INPUT_FONT
    ws["B3"].fill = INPUT_FILL

    sections = {
        5: ("TRUCK COST PARAMETERS", [
            ("Truck Capacity (bikes)", COST_MODEL["truck_capacity_bikes"], ""),
            ("Fixed Cost per Truck/Night ($)", COST_MODEL["fixed_cost_per_truck_per_night_usd"], CURRENCY_FMT),
            ("Variable Cost per Mile ($)", COST_MODEL["variable_cost_per_mile_usd"], CURRENCY_FMT),
            ("Driver Hourly Wage ($)", COST_MODEL["driver_hourly_wage_usd"], CURRENCY_FMT),
            ("Overtime Multiplier", COST_MODEL["overtime_multiplier"], "0.0x"),
            ("Overtime Threshold (hrs)", COST_MODEL["overtime_threshold_hours"], "0.0"),
            ("Fuel Cost per Gallon ($)", COST_MODEL["fuel_cost_per_gallon_usd"], CURRENCY_FMT),
            ("Truck MPG", COST_MODEL["truck_mpg"], "0.0"),
            ("Load/Unload Time per Stop (min)", COST_MODEL["avg_load_unload_minutes_per_stop"], "0"),
            ("Avg Overnight Speed (mph)", COST_MODEL["avg_speed_mph_overnight"], "0"),
        ]),
        17: ("REVENUE MODEL", [
            ("Revenue per Trip ($)", REVENUE_MODEL["revenue_per_trip_usd"], CURRENCY_FMT),
            ("Annual Membership Daily Equiv ($)", REVENUE_MODEL["annual_membership_daily_equiv_usd"], CURRENCY_FMT),
            ("Single Ride Price ($)", REVENUE_MODEL["single_ride_price_usd"], CURRENCY_FMT),
            ("Day Pass Price ($)", REVENUE_MODEL["day_pass_price_usd"], CURRENCY_FMT),
            ("E-bike Unlock Fee ($)", REVENUE_MODEL["ebike_unlock_fee_usd"], CURRENCY_FMT),
        ]),
        24: ("OPPORTUNITY COST", [
            ("Lost Trip Revenue ($)", OPPORTUNITY_COST["lost_trip_revenue_usd"], CURRENCY_FMT),
            ("Customer Dissatisfaction Penalty ($)", OPPORTUNITY_COST["customer_dissatisfaction_penalty_usd"], CURRENCY_FMT),
            ("Brand Damage Multiplier", OPPORTUNITY_COST["brand_damage_multiplier"], "0.00x"),
        ]),
        29: ("TARGET FILL RATIOS", [
            ("Morning Min Fill", TARGET_FILL_RATIO["morning_min"], PCT_FMT),
            ("Morning Max Fill", TARGET_FILL_RATIO["morning_max"], PCT_FMT),
            ("Ideal Fill", TARGET_FILL_RATIO["ideal"], PCT_FMT),
        ]),
        34: ("SIMULATION PARAMETERS", [
            ("Number of Trials", SIMULATION["n_trials"], NUM_FMT),
            ("Random Seed", SIMULATION["random_seed"], "0"),
        ]),
    }

    for start_row, (title, params) in sections.items():
        ws.cell(row=start_row, column=1, value=title).font = SUBHEADER_FONT
        ws.cell(row=start_row, column=1).fill = SUBHEADER_FILL
        ws.cell(row=start_row, column=2).fill = SUBHEADER_FILL

        for i, (label, value, fmt) in enumerate(params):
            r = start_row + 1 + i
            ws.cell(row=r, column=1, value=label).font = DATA_FONT
            cell = ws.cell(row=r, column=2, value=value)
            cell.font = INPUT_FONT
            cell.fill = INPUT_FILL
            if fmt:
                cell.number_format = fmt

    ws["A38"] = "Blue = editable inputs. Modify these to run scenarios."
    ws["A38"].font = Font(name="Arial", size=9, italic=True, color="666666")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20


def build_imbalance_sheet(wb):
    ws = wb.create_sheet("Imbalance")
    ws.sheet_properties.tabColor = "639922"

    ws["A1"] = "Midnight Inventory vs. Morning Target"
    ws["A1"].font = TITLE_FONT

    targets = pd.read_parquet(DATA_PROCESSED / "rebalancing_targets.parquet")

    headers = ["Station", "Capacity", "Midnight Fill", "Midnight %",
               "Target Bikes", "Target %", "Deficit", "Needs Pickup", "Needs Dropoff"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))

    for i, (_, row) in enumerate(targets.iterrows()):
        r = i + 4
        ws.cell(row=r, column=1, value=row["name"]).font = DATA_FONT
        ws.cell(row=r, column=2, value=int(row["capacity"])).font = DATA_FONT
        ws.cell(row=r, column=3, value=int(row["estimated_midnight_fill"])).font = DATA_FONT
        ws.cell(row=r, column=4, value=f"=C{r}/B{r}").font = FORMULA_FONT
        ws.cell(row=r, column=4).number_format = PCT_FMT
        ws.cell(row=r, column=5, value=int(row["target_bikes"])).font = DATA_FONT
        ws.cell(row=r, column=6, value=f"=E{r}/B{r}").font = FORMULA_FONT
        ws.cell(row=r, column=6).number_format = PCT_FMT
        ws.cell(row=r, column=7, value=f"=E{r}-C{r}").font = FORMULA_FONT
        ws.cell(row=r, column=8, value=f'=IF(G{r}<0,-G{r},0)').font = FORMULA_FONT
        ws.cell(row=r, column=9, value=f'=IF(G{r}>0,G{r},0)').font = FORMULA_FONT

        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = THIN_BORDER

    last_data = len(targets) + 4
    ws.cell(row=last_data, column=1, value="TOTALS").font = SUBHEADER_FONT
    for col_idx in [2, 3, 5, 8, 9]:
        cl = get_column_letter(col_idx)
        ws.cell(row=last_data, column=col_idx, value=f"=SUM({cl}4:{cl}{last_data-1})").font = FORMULA_FONT

    r = last_data + 2
    ws.cell(row=r, column=1, value="COST OF NO REBALANCING").font = SUBHEADER_FONT
    ws.cell(row=r, column=1).fill = SUBHEADER_FILL

    ws.cell(row=r+1, column=1, value="Unmet Trips / Morning").font = DATA_FONT
    ws.cell(row=r+1, column=2, value=f"=H{last_data}+I{last_data}").font = FORMULA_FONT

    ws.cell(row=r+2, column=1, value="Daily Lost Revenue ($)").font = DATA_FONT
    ws.cell(row=r+2, column=2, value=f"=B{r+1}*'Cost Model'!B25").font = LINK_FONT
    ws.cell(row=r+2, column=2).number_format = CURRENCY_FMT

    ws.cell(row=r+3, column=1, value="Daily Dissatisfaction ($)").font = DATA_FONT
    ws.cell(row=r+3, column=2, value=f"=B{r+1}*'Cost Model'!B26").font = LINK_FONT
    ws.cell(row=r+3, column=2).number_format = CURRENCY_FMT

    ws.cell(row=r+4, column=1, value="Daily Total Opportunity Cost ($)").font = SUBHEADER_FONT
    ws.cell(row=r+4, column=2, value=f"=(B{r+2}+B{r+3})*'Cost Model'!B27").font = LINK_FONT
    ws.cell(row=r+4, column=2).number_format = CURRENCY_FMT

    ws.cell(row=r+5, column=1, value="Monthly Opportunity Cost ($)").font = SUBHEADER_FONT
    ws.cell(row=r+5, column=2, value=f"=B{r+4}*30").font = FORMULA_FONT
    ws.cell(row=r+5, column=2).number_format = CURRENCY_INT

    auto_width(ws)


def build_optimization_sheet(wb):
    ws = wb.create_sheet("Optimization")
    ws.sheet_properties.tabColor = "FF8C00"

    ws["A1"] = "Truck Routing Optimization"
    ws["A1"].font = TITLE_FONT

    opt_sweep = pd.read_parquet(DATA_PROCESSED / "optimization_sweep.parquet")
    trucks = pd.read_parquet(DATA_PROCESSED / "optimal_routes.parquet")

    ws["A3"] = "Optimization Sweep (1–8 Trucks)"
    ws["A3"].font = SUBHEADER_FONT
    ws["A3"].fill = SUBHEADER_FILL

    opt_headers = ["Trucks", "Total Cost ($)", "Total Miles", "Bikes Moved", "Service Level"]
    for col, h in enumerate(opt_headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(opt_headers))

    for i, (_, row) in enumerate(opt_sweep.iterrows()):
        r = i + 5
        ws.cell(row=r, column=1, value=int(row["n_trucks"])).font = DATA_FONT
        ws.cell(row=r, column=2, value=round(row["total_cost"], 2)).font = DATA_FONT
        ws.cell(row=r, column=2).number_format = CURRENCY_FMT
        ws.cell(row=r, column=3, value=round(row["total_miles"], 1)).font = DATA_FONT
        ws.cell(row=r, column=4, value=int(row["total_bikes"])).font = DATA_FONT
        ws.cell(row=r, column=5, value=round(row["service_level"], 3)).font = DATA_FONT
        ws.cell(row=r, column=5).number_format = PCT_FMT
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = THIN_BORDER

    ws["A15"] = "Optimal Route Details (4 Trucks)"
    ws["A15"].font = SUBHEADER_FONT
    ws["A15"].fill = SUBHEADER_FILL

    truck_headers = ["Truck #", "Stops", "Miles", "Drive Hrs", "Stop Hrs",
                     "Total Hrs", "Bikes Moved", "Fixed ($)", "Mileage ($)",
                     "Fuel ($)", "Labor ($)", "Total Cost ($)"]
    for col, h in enumerate(truck_headers, 1):
        ws.cell(row=16, column=col, value=h)
    style_header_row(ws, 16, len(truck_headers))

    for i, (_, row) in enumerate(trucks.iterrows()):
        r = i + 17
        ws.cell(row=r, column=1, value=int(row.get("truck_id", i+1))).font = DATA_FONT
        ws.cell(row=r, column=2, value=int(row.get("stops", 0))).font = DATA_FONT
        ws.cell(row=r, column=3, value=round(row.get("miles", 0), 1)).font = DATA_FONT
        ws.cell(row=r, column=4, value=round(row.get("drive_hours", 0), 2)).font = DATA_FONT
        ws.cell(row=r, column=5, value=round(row.get("stop_hours", 0), 2)).font = DATA_FONT
        ws.cell(row=r, column=6, value=round(row.get("total_hours", 0), 2)).font = DATA_FONT
        ws.cell(row=r, column=7, value=int(row.get("bikes_moved", 0))).font = DATA_FONT
        ws.cell(row=r, column=8, value=round(row.get("fixed_cost", 0), 2)).font = DATA_FONT
        ws.cell(row=r, column=8).number_format = CURRENCY_FMT
        ws.cell(row=r, column=9, value=round(row.get("mileage_cost", 0), 2)).font = DATA_FONT
        ws.cell(row=r, column=9).number_format = CURRENCY_FMT
        ws.cell(row=r, column=10, value=round(row.get("fuel_cost", 0), 2)).font = DATA_FONT
        ws.cell(row=r, column=10).number_format = CURRENCY_FMT
        ws.cell(row=r, column=11, value=round(row.get("labor_cost", 0), 2)).font = DATA_FONT
        ws.cell(row=r, column=11).number_format = CURRENCY_FMT
        ws.cell(row=r, column=12).font = FORMULA_FONT
        ws.cell(row=r, column=12).number_format = CURRENCY_FMT
        ws.cell(row=r, column=12, value=f"=SUM(H{r}:K{r})")

        for col in range(1, len(truck_headers) + 1):
            ws.cell(row=r, column=col).border = THIN_BORDER

    auto_width(ws)


def build_simulation_sheet(wb):
    ws = wb.create_sheet("Simulation")
    ws.sheet_properties.tabColor = "9B59B6"

    ws["A1"] = "Monte Carlo Simulation Results"
    ws["A1"].font = TITLE_FONT

    summary = pd.read_csv(DATA_PROCESSED / "simulation_summary.csv")

    ws["A3"] = "SIMULATION CONFIGURATION"
    ws["A3"].font = SUBHEADER_FONT
    ws["A3"].fill = SUBHEADER_FILL

    ws.cell(row=4, column=1, value="Trucks Deployed").font = DATA_FONT
    ws.cell(row=4, column=2, value=4).font = DATA_FONT

    ws.cell(row=5, column=1, value="Number of Trials").font = DATA_FONT
    ws.cell(row=5, column=2, value=f"='Cost Model'!B35").font = LINK_FONT

    ws.cell(row=6, column=1, value="Random Seed").font = DATA_FONT
    ws.cell(row=6, column=2, value=f"='Cost Model'!B36").font = LINK_FONT

    ws["A8"] = "CRYSTAL BALL ASSUMPTIONS (Random Variables)"
    ws["A8"].font = SUBHEADER_FONT
    ws["A8"].fill = SUBHEADER_FILL

    assumptions = [
        ("Morning Demand (per station)", "Poisson(λ) or Normal(μ,σ)", "Fitted from 12 months of data"),
        ("Weather Factor", "Discrete: 1.0 (65%), 0.85 (15%), 0.6 (10%), 1.1 (10%)", "Historical weather pattern"),
        ("Day Type", "Bernoulli: Weekday 71.4%, Weekend 28.6%", "Calendar distribution"),
    ]
    headers_a = ["Assumption", "Distribution", "Source"]
    for col, h in enumerate(headers_a, 1):
        ws.cell(row=9, column=col, value=h)
    style_header_row(ws, 9, 3)

    for i, (name, dist, src) in enumerate(assumptions):
        r = 10 + i
        ws.cell(row=r, column=1, value=name).font = DATA_FONT
        ws.cell(row=r, column=2, value=dist).font = DATA_FONT
        ws.cell(row=r, column=3, value=src).font = DATA_FONT
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER

    ws["A14"] = "FORECAST RESULTS"
    ws["A14"].font = SUBHEADER_FONT
    ws["A14"].fill = SUBHEADER_FILL

    result_headers = ["Metric", "Mean", "P10", "P25", "P50", "P75", "P90"]
    for col, h in enumerate(result_headers, 1):
        ws.cell(row=15, column=col, value=h)
    style_header_row(ws, 15, len(result_headers))

    s = summary.iloc[0]

    metrics = [
        ("Service Level", s.get("mean_service_level", 0), s.get("p10_service_level", 0),
         s.get("p25_service_level", 0) if "p25_service_level" in s else "",
         s.get("p50_service_level", 0), "",
         s.get("p90_service_level", 0)),
        ("Unmet Trips", s.get("mean_unmet_trips", 0), "", "", "", "", s.get("p90_unmet_trips", 0)),
        ("Opportunity Cost ($)", s.get("mean_opportunity_cost", 0), "", "", "", "", ""),
        ("Rebalancing Cost ($)", s.get("rebalancing_cost", 0), "", "", "", "", ""),
        ("Net Benefit ($)", s.get("mean_net_benefit", 0), s.get("p10_net_benefit", 0),
         "", s.get("p50_net_benefit", 0), "", s.get("p90_net_benefit", 0)),
    ]

    for i, metric_row in enumerate(metrics):
        r = 16 + i
        ws.cell(row=r, column=1, value=metric_row[0]).font = DATA_FONT
        for col_idx in range(1, len(metric_row)):
            val = metric_row[col_idx]
            if val == "":
                continue
            cell = ws.cell(row=r, column=col_idx + 1, value=round(float(val), 2) if isinstance(val, (int, float, np.floating)) else val)
            cell.font = DATA_FONT
            if "Service" in metric_row[0]:
                cell.number_format = PCT_FMT
            elif "$" in metric_row[0]:
                cell.number_format = CURRENCY_FMT
            elif "Trips" in metric_row[0]:
                cell.number_format = NUM_FMT
            cell.border = THIN_BORDER

    ws["A23"] = "CONFIDENCE INTERVAL"
    ws["A23"].font = SUBHEADER_FONT
    ws["A23"].fill = SUBHEADER_FILL

    ws.cell(row=24, column=1, value="P(Net Benefit > 0)").font = DATA_FONT
    ws.cell(row=24, column=2, value=round(float(s.get("prob_positive_net_benefit", 0)), 3)).font = DATA_FONT
    ws.cell(row=24, column=2).number_format = PCT_FMT

    auto_width(ws)


def build_sensitivity_sheet(wb):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_properties.tabColor = "27AE60"

    ws["A1"] = "Sensitivity Analysis & Efficient Frontier"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "TRUCK COUNT SWEEP (SolverTable)"
    ws["A3"].font = SUBHEADER_FONT
    ws["A3"].fill = SUBHEADER_FILL

    sweep = pd.read_parquet(DATA_PROCESSED / "sensitivity_truck_sweep.parquet")

    sweep_headers = ["Trucks", "Rebal Cost ($)", "Mean Service Level",
                     "P10 Service Level", "P90 Service Level",
                     "Mean Net Benefit ($/night)", "P(Benefit > 0)"]
    for col, h in enumerate(sweep_headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(sweep_headers))

    for i, (_, row) in enumerate(sweep.iterrows()):
        r = i + 4
        ws.cell(row=r, column=1, value=int(row["n_trucks"])).font = DATA_FONT
        ws.cell(row=r, column=2, value=round(row["rebal_cost"], 2)).font = DATA_FONT
        ws.cell(row=r, column=2).number_format = CURRENCY_FMT
        ws.cell(row=r, column=3, value=round(row["mean_service_level"], 3)).font = DATA_FONT
        ws.cell(row=r, column=3).number_format = PCT_FMT
        ws.cell(row=r, column=4, value=round(row["p10_service_level"], 3)).font = DATA_FONT
        ws.cell(row=r, column=4).number_format = PCT_FMT
        ws.cell(row=r, column=5, value=round(row["p90_service_level"], 3)).font = DATA_FONT
        ws.cell(row=r, column=5).number_format = PCT_FMT
        ws.cell(row=r, column=6, value=round(row["mean_net_benefit"], 2)).font = DATA_FONT
        ws.cell(row=r, column=6).number_format = CURRENCY_FMT
        ws.cell(row=r, column=7, value=round(row["prob_positive"], 3)).font = DATA_FONT
        ws.cell(row=r, column=7).number_format = PCT_FMT
        for col in range(1, len(sweep_headers) + 1):
            ws.cell(row=r, column=col).border = THIN_BORDER

    r = len(sweep) + 5
    ws.cell(row=r, column=1, value="PARAMETER SENSITIVITY").font = SUBHEADER_FONT
    ws.cell(row=r, column=1).fill = SUBHEADER_FILL

    cap_data = pd.read_csv(DATA_PROCESSED / "sensitivity_truck_capacity.csv")
    ws.cell(row=r+1, column=1, value="Truck Capacity").font = SUBHEADER_FONT
    cap_headers = ["Capacity (bikes)", "Cost ($)", "Service Level"]
    for col, h in enumerate(cap_headers, 1):
        ws.cell(row=r+2, column=col, value=h)
    style_header_row(ws, r+2, 3)

    for i, (_, crow) in enumerate(cap_data.iterrows()):
        rr = r + 3 + i
        ws.cell(row=rr, column=1, value=int(crow["truck_capacity"])).font = DATA_FONT
        ws.cell(row=rr, column=2, value=round(crow["cost"], 2)).font = DATA_FONT
        ws.cell(row=rr, column=2).number_format = CURRENCY_FMT
        ws.cell(row=rr, column=3, value=round(crow["service_level"], 3)).font = DATA_FONT
        ws.cell(row=rr, column=3).number_format = PCT_FMT

    wage_start = r + 3 + len(cap_data) + 1
    wage_data = pd.read_csv(DATA_PROCESSED / "sensitivity_driver_wage.csv")
    ws.cell(row=wage_start, column=1, value="Driver Wage").font = SUBHEADER_FONT
    for col, h in enumerate(["Wage ($/hr)", "Total Cost ($)"], 1):
        ws.cell(row=wage_start+1, column=col, value=h)
    style_header_row(ws, wage_start+1, 2)

    for i, (_, wrow) in enumerate(wage_data.iterrows()):
        rr = wage_start + 2 + i
        ws.cell(row=rr, column=1, value=wrow["driver_wage"]).font = DATA_FONT
        ws.cell(row=rr, column=1).number_format = CURRENCY_FMT
        ws.cell(row=rr, column=2, value=round(wrow["cost"], 2)).font = DATA_FONT
        ws.cell(row=rr, column=2).number_format = CURRENCY_FMT

    auto_width(ws)


def build_figures_sheet(wb):
    ws = wb.create_sheet("Figures")
    ws.sheet_properties.tabColor = "34495E"

    ws["A1"] = "Generated Figures"
    ws["A1"].font = TITLE_FONT

    from openpyxl.drawing.image import Image as XlImage
    from config import FIGURES_DIR as FIG_DIR
    import os

    figure_files = [
        ("01_demand_overview.png", "Demand Profile Overview"),
        ("02_station_distributions.png", "Station Demand Distributions"),
        ("03_imbalance_analysis.png", "Imbalance Analysis"),
        ("04_optimal_routes.png", "Optimal Truck Routes"),
        ("05_simulation_results.png", "Monte Carlo Simulation Results"),
        ("06_efficient_frontier.png", "Efficient Frontier"),
        ("07_sensitivity_tornado.png", "Sensitivity Tornado"),
        ("08_cost_of_inaction.png", "Cost of Inaction Analysis"),
        ("09_operational_costs.png", "Operational Cost Analysis"),
    ]

    row = 3
    for fname, title in figure_files:
        fpath = FIG_DIR / fname
        ws.cell(row=row, column=1, value=title).font = SUBHEADER_FONT
        if fpath.exists():
            try:
                img = XlImage(str(fpath))
                img.width = 720
                img.height = int(img.height * (720 / img.width)) if img.width > 0 else 400
                ws.add_image(img, f"A{row + 1}")
                row += max(20, img.height // 15) + 2
            except Exception:
                ws.cell(row=row + 1, column=1, value=f"[Image: {fname}]").font = DATA_FONT
                row += 3
        else:
            ws.cell(row=row + 1, column=1, value=f"[Not generated: {fname}]").font = DATA_FONT
            row += 3


def main():
    wb = Workbook()

    build_summary_sheet(wb)
    build_station_data_sheet(wb)
    build_cost_model_sheet(wb)
    build_imbalance_sheet(wb)
    build_optimization_sheet(wb)
    build_simulation_sheet(wb)
    build_sensitivity_sheet(wb)
    build_figures_sheet(wb)

    output_path = EXCEL_DIR / "citibike_rebalancing.xlsx"
    wb.save(output_path)
    print(f"[saved] {output_path}")
    print(f"  Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
